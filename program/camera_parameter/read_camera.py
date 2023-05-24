import os
import numpy as np
import pandas as pd
import openpyxl
from scipy.interpolate import interp1d


#####################
# read_cam
# input: file_name[str]
# output:   result.quantum_efficiency[np.array], result.transparency[np.array], result.wavelength_set[nm],
#           result.camera_total_efficiency[dict]
# dimension: camera_total_efficiency[136, 9], 136: wavelength number, 9: channel_number
#####################
def read_cam(QE_address, T_address):
    QE_raw = pd.read_excel(QE_address, 'QE')
    T_raw = pd.read_excel(T_address)
    result = {}
    camera_total_efficiency = {}
    result['quantum_efficiency'] = QE_raw.to_numpy()[:, 1::]
    wavelength_set = QE_raw.to_numpy()[:, 0]
    t_interp = interp1d(T_raw.to_numpy()[:, 0]*1000, T_raw.to_numpy()[:, 1], kind='linear', fill_value='extrapolate')
    result['transparency'] = t_interp(wavelength_set)
    result['wavelength_set'] = wavelength_set
    wl, channel = np.shape(QE_raw.to_numpy()[:, 1::])
    for i in range(channel):
        camera_total_efficiency['channel_' + str(i)] = {}
        for j in range(wl):
            camera_total_efficiency['channel_' + str(i)][str(int(wavelength_set[j]))] = result['quantum_efficiency'][j, i] * result['transparency'][j]
    result['camera_total_efficiency'] = camera_total_efficiency
    return result


##############################################
# save file in to .xls
##############################################
def save_file(t_field, real_radiation, digital_radiation, temperature_center, emissivity_set):
    dir_name = 'T' + str(temperature_center) + '_' + str(emissivity_set) + '_digital'
    if not os.path.exists(os.path.join('data', dir_name)):
        os.mkdir(os.path.join('data', dir_name))

    # real_radiation
    file_radiation = os.path.join('data', dir_name, 'real_radiation_' + str(temperature_center)  + '.xlsx')
    workbook_radiation = openpyxl.Workbook()
    worksheet_radiation = workbook_radiation.active
    for wl in real_radiation:
        worksheet_radiation = workbook_radiation.create_sheet(wl)
        for row in real_radiation[wl]:
            worksheet_radiation.append(list(row))
    workbook_radiation.remove(workbook_radiation['Sheet'])
    workbook_radiation.save(file_radiation)

    # t_field
    file_t = os.path.join('data', dir_name, 't_field_' + str(temperature_center)  + '.xlsx')
    workbook_t = openpyxl.Workbook()
    worksheet_t = workbook_t.active

    for row in t_field:
        worksheet_t.append(list(row))
    workbook_t.save(file_t)

    # digital_radiation
    file_digit = os.path.join('data', dir_name, 'digital_value_' + str(temperature_center)  + '.xlsx')
    workbook_digit = openpyxl.Workbook()
    worksheet_digit = workbook_digit.active
    for channel in digital_radiation:
        worksheet_digit = workbook_digit.create_sheet(channel)
        for row in digital_radiation[channel]:
            worksheet_digit.append(list(row))
    workbook_digit.remove(workbook_digit['Sheet'])
    workbook_digit.save(file_digit)
