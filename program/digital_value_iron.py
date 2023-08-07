import os
import multiprocessing as mp
import numpy as np
import math
import warnings
import openpyxl
import time
import matplotlib.pyplot as plt

import pandas as pd
from scipy.interpolate import interp1d
from scipy.integrate import quad
from joblib import Parallel, delayed
# from PIL import Image


def digital_value_rebuild(emissivity_set):
    ###############################################
    # parameter definition
    ###############################################
    print("Number of processors: ", mp.cpu_count())
    image_resolution = [50, 50]                                               # [pixel] [100 * 100]
    diameter_ratio = 0.9                                                        # adjust the visualisation of data_field
    shutter_time = 200                                                          # adjust explosure time
    temperature_center = 3500                                                 # temperature of the center_area
    temperature_background = 2500                                                # set background temperature to 50K as black body
    melt_temperature = 1811                                                     # set melt temperature
    emissivity_liquid = 0.07                                                    # set emissivity in liquid phase
    temperature_distribution = 'linear'                                         # gaussian / linear / sigmoid
    # emissivity_set = 34                                                          # which data set is used, [0] stands for black body radiation

    # explosure_time = 200                                                        # explosure time of camera
    plot_channel = 'channel_3'                                                      # set plot configuration

    # read camera parameter
    qe_address = os.path.join('camera_parameter', 'CMS22010236.xlsx')
    tr_address = os.path.join('camera_parameter', 'FIFO-Lens_tr.xls')
    cam_param = read_cam(qe_address, tr_address)
    cam_efficiency = cam_param['camera_total_efficiency']
    # print(cam_efficiency)

    # generate temperature field
    t_field = temperature_field(image_resolution, diameter_ratio, temperature_center,
                                            temperature_background, temperature_distribution)
    # print(t_field)

    # read emissivity data and interpolation
    # original emissivity data from iron
    emissivity_liquid_address = os.path.join("hypothetical", "emissivity_") + "iron_liquid" + ".txt"
    emissivity_solid_address = os.path.join("hypothetical", "emissivity_") + "iron_solid" + ".txt"
    emissivity_temp_address = os.path.join("hypothetical", "emissivity_") + "iron_temp" + ".txt"

    data_liquid_raw = np.loadtxt(emissivity_liquid_address)
    data_solid_raw = np.loadtxt(emissivity_solid_address)
    data_temp_raw = np.loadtxt(emissivity_temp_address)
    data_temp_raw[:, 1] = data_temp_raw[:, 1] / data_temp_raw[16, 1]

    interp_liquid_emissivity = interp1d(data_liquid_raw[:, 0], data_liquid_raw[:, 1], kind='linear',
                                        fill_value='extrapolate')
    interp_solid_emissivity = interp1d(data_solid_raw[:, 0], data_solid_raw[:, 1], kind='linear',
                                        fill_value='extrapolate')

    final_results = {}
    # start calculation in channel
    for channel in cam_efficiency.keys():
        cam_efficiency_channel = [[int(k), v] for k, v in cam_efficiency[channel].items()]
        cam_efficiency_function = interp1d([i[0] for i in cam_efficiency_channel], [i[1] for i in cam_efficiency_channel], kind='linear', fill_value='extrapolate')

        def radiation_integration(temperature, emissivity_solid, emissivity_liquid, cam_efficiency_data, data_temp):
            def radiation(wl):
                return black_body_radiation(temperature, wl) * emissivity(temperature, wl, emissivity_solid, emissivity_liquid, melt_temperature, data_temp) * camera_efficiency(wl, cam_efficiency_data)

            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                radiation_rec = quad(radiation, 500 * (10**-9), 900 * (10**-9), limit=5)[0] * shutter_time      # 200 means the calibration factor and shutter time
            return radiation_rec
        t_field_reshape = t_field.reshape(t_field.shape[0] * t_field.shape[1])
        digital_value = Parallel(n_jobs=mp.cpu_count()-1)(delayed(radiation_integration)(t_field_reshape[i], interp_solid_emissivity, interp_liquid_emissivity, cam_efficiency_function, data_temp_raw) for i in range(len(t_field_reshape)))
        final_results[channel] = np.array(digital_value).reshape(image_resolution).astype(int)

    # calculate emissivity_map
    emi_field = emissivity_map(t_field, interp_solid_emissivity, interp_liquid_emissivity, melt_temperature, data_temp_raw)

    # print(final_results)
    save_file(t_field, final_results, emi_field, temperature_center, emissivity_set)
    emi_plot(interp_solid_emissivity, interp_liquid_emissivity, melt_temperature, data_temp_raw)
    # plot figures
    fig, axs = plt.subplots(3, 1)
    y_radiation = final_results[plot_channel][len(final_results[plot_channel])//2]
    y_t         = t_field[len(t_field)//2]
    y_emi       = emi_field[len(emi_field)//2]
    x_radiation = np.arange(len(y_radiation))
    axs[0].plot(x_radiation, y_radiation)
    axs[1].plot(x_radiation, y_t)
    axs[2].plot(x_radiation, y_emi)
    axs[0].set(title='digital value' + plot_channel, ylabel='digital_value')
    axs[1].set(title='temperature', ylabel='temperature')
    axs[2].set(title='emissivity', xlabel='position', ylabel='emissivity')
    plt.tight_layout()
    # plt.show()


###############################################
# define function
###############################################
def black_body_radiation(temperature, wavelength):
    h = 6.62607015e-34      # Plank's constant
    c = 299792458           # Speed of light
    k_b = 1.380649e-23      # Boltzmann's constant
    radiance = 2 * h * np.float_power(c, 2) * np.float_power(wavelength, -5) / (math.exp(h*c/(k_b * wavelength * temperature))-1)
    return radiance


def emi_plot(emissivity_solid, emissivity_liquid, melt_temperature, data_temp):
    t_field = np.linspace(1500, 2000, 30)
    wl_set = np.linspace(400, 800, 30) * 1e-9
    T, WL = np.meshgrid(t_field, wl_set)
    emi = np.zeros(T.shape)
    for i in range(len(t_field)):
        for j in range(len(wl_set)):
            emi[i, j] = emissivity(T[i, j], WL[i, j], emissivity_solid, emissivity_liquid, melt_temperature, data_temp)
    fig = plt.figure(figsize=(8, 6), dpi=800)
    ax = fig.add_subplot(111, projection='3d')
    # 绘制散点图
    ax.scatter(T, WL * 1e9, emi, c='g')

    # 设置坐标轴标签
    ax.set_xlabel('Temperature[K]')
    ax.set_ylabel('Wavelength[nm]')
    ax.set_zlabel('Emissivity')

    plt.savefig('emissivity_model.jpg')
    # plt.show()
    return 0


def factor_temperature(temperature, data_temp):
    factor_fun = interp1d(data_temp[:, 0], data_temp[:, 1], kind="linear", fill_value='extrapolate')
    return factor_fun(temperature)


def emissivity(temperature, wavelength, emissivity_solid, emissivity_liquid, melt_temperature, data_temp):
    fact_temp = factor_temperature(temperature, data_temp)
    if temperature<melt_temperature:
        emi = min(emissivity_solid(wavelength * (10 ** 9)) * fact_temp, 1)
    else:
        emi = min(emissivity_liquid(wavelength * (10 ** 9)) * fact_temp, 1)
    return emi


def camera_efficiency(wavelength, cam_fun):
    qe = cam_fun(wavelength * (10**9))
    return qe


def emissivity_map(t_field, emissivity_solid, emissivity_liquid, melt_temperature, data_temp):
    row, colum = t_field.shape
    emi_field = np.zeros(t_field.shape)
    wavelength = np.arange(500 * (10**-9), 901 * (10**-9))
    for i in range(row):
        for j in range(colum):
            emi_field[i, j] = np.average([emissivity(t_field[i, j], wl, emissivity_solid, emissivity_liquid, melt_temperature, data_temp) for wl in wavelength])
    return emi_field


def save_file(t_field, digital_radiation, emi_field, temperature_center, emissivity_set):
    emissivity_set = emissivity_set

    dir_name = 'T' + str(temperature_center) + '_' + str(emissivity_set) + '_digital'
    if not os.path.exists(os.path.join('data', dir_name)):
        os.mkdir(os.path.join('data', dir_name))

    # t_field
    file_t = os.path.join('data', dir_name, 't_field_' + str(temperature_center) + '.xlsx')
    workbook_t = openpyxl.Workbook()
    worksheet_t = workbook_t.active
    # image_t = Image.fromarray(t_field).convert("L")
    # image_t.save(os.path.join('data', dir_name, 't_field' + '.png'))
    plt.imshow(t_field, cmap='viridis')
    plt.colorbar()
    plt.xlabel('X_position')
    plt.ylabel('Y_position')
    plt.title('Temperature_map')
    plt.savefig(os.path.join('data', dir_name, 't_field' + '.jpg'))
    plt.clf()

    for row in t_field:
        worksheet_t.append(list(row))
    workbook_t.save(file_t)

    # emi_field
    file_emi = os.path.join('data', dir_name, 'emi_field_' + str(temperature_center) + '.xlsx')
    workbook_emi = openpyxl.Workbook()
    worksheet_emi = workbook_emi.active
    # image_emi = Image.fromarray(emi_field).convert("L")
    # image_emi.save(os.path.join('data', dir_name, 'emi_field' + '.png'))
    plt.imshow(emi_field, cmap='viridis')
    plt.colorbar()
    plt.xlabel('X_position')
    plt.ylabel('Y_position')
    plt.title('Emissivity_map')
    plt.savefig(os.path.join('data', dir_name, 'emi_field' + '.jpg'))
    plt.clf()

    for row in emi_field:
        worksheet_emi.append(list(row))
    workbook_emi.save(file_emi)

    # digital_radiation
    file_digit = os.path.join('data', dir_name, 'digital_value_' + str(temperature_center)  + '.xlsx')
    workbook_digit = openpyxl.Workbook()
    worksheet_digit = workbook_digit.active
    find_edge_value = (lambda x: (np.amin(x), np.amax(x)))
    v_min, v_max = find_edge_value(np.vstack([digital_radiation[channel] for channel in digital_radiation.keys()]))

    for channel in digital_radiation:
        worksheet_digit = workbook_digit.create_sheet(channel)
        for row in digital_radiation[channel]:
            worksheet_digit.append(list(row))
        plt.imshow(digital_radiation[channel], cmap='viridis', vmin=v_min, vmax=v_max)
        plt.colorbar()
        plt.xlabel('X_position')
        plt.ylabel('Y_position')
        plt.title('Digital_value_' + channel)
        plt.savefig(os.path.join('data', dir_name, 'digital_value_' + channel + '.jpg'))
        plt.clf()
    workbook_digit.remove(workbook_digit['Sheet'])
    workbook_digit.save(file_digit)


def read_cam(QE_address, T_address):
    QE_raw = pd.read_excel(QE_address, 'QE')
    T_raw = pd.read_excel(T_address)
    result = {}
    camera_total_efficiency = {}
    result['quantum_efficiency'] = QE_raw.to_numpy()[:, 1::]
    wavelength_set = QE_raw.to_numpy()[:, 0]
    t_interp = interp1d(T_raw.to_numpy()[:, 0]*1000, T_raw.to_numpy()[:, 1], kind='linear', fill_value='extrapolate')
    result['transparency'] = t_interp(wavelength_set)
    result['wavelength_set'] = wavelength_set
    wl, channel = np.shape(QE_raw.to_numpy()[:, 1::])
    for i in range(channel):
        camera_total_efficiency['channel_' + str(i)] = {}
        for j in range(wl):
            camera_total_efficiency['channel_' + str(i)][str(int(wavelength_set[j]))] = result['quantum_efficiency'][j, i] * result['transparency'][j]
    result['camera_total_efficiency'] = camera_total_efficiency
    return result


def temperature_field(resolution, diameter_ratio, temperature_center, temperature_background, distribution_type):
    # calculation of geometric relations
    row_center = resolution[0] / 2
    colum_center = resolution[1] / 2
    radius = round(min(resolution[0], resolution[1]) * diameter_ratio / 2)
    sigma = radius * 2.5

    # initialise the temperature field
    field = np.ones(resolution)

    # start calculation
    if distribution_type == 'sigmoid':
        for i in range(resolution[0]):
            for j in range(resolution[1]):
                field[i, j] = round((temperature_background + (temperature_center - temperature_background) *
                                    sigmoid((-((i - row_center)**2 + (j - colum_center)**2) + radius**2) /
                                            radius**2 * 1000)))
    elif distribution_type == 'linear':
        for i in range(resolution[0]):
            for j in range(resolution[1]):
                field[i, j] = max(round(temperature_center + (temperature_background - temperature_center) *
                                        np.sqrt((i - row_center)**2 + (j - colum_center)**2) / radius), temperature_background)
    elif distribution_type == 'gaussian':
        for i in range(resolution[0]):
            for j in range(resolution[1]):
                field[i, j] = max(round(temperature_background + (temperature_center - temperature_background) *
                                        distribution(((i - row_center)**2 + (j - colum_center)**2), sigma)),
                                  temperature_background)
    return field


def sigmoid(x):
    return 1. / (1. + math.exp(-x/100))


def distribution(x, sigma):
    result = 1 / math.exp(x / (2 * sigma**2))
    return result


if 1:
    start_time = time.perf_counter()

    emissivity_set = 5
    digital_value_rebuild(emissivity_set)


    end_time = time.perf_counter()
    print("calculation time: ", end_time - start_time, " second")