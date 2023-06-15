import os
import re
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import time
from scipy.ndimage.filters import gaussian_filter
import math
import warnings
import openpyxl
import multiprocessing as mp
import tifffile
from scipy.constants import c, h, k
from scipy.optimize import curve_fit, least_squares
from scipy.integrate import quad
from scipy.interpolate import interp1d
from joblib import Parallel, delayed


def t_estimate(data_path, temperature_center):
    # file path setting
    # data_path = 'V0_bb_750'
    emissivity_set = '0'
    result_dir = 'result_v050_gray_body'
    # center_temperature = 750

    # read intensity data
    intensity_digital_value, exposure_time = image_read(data_path)

    # read camera parameter
    camera_folder = 'camera_parameter'
    sensor_param_tr, sensor_param_qe = camera_read(camera_folder)

    # wavelength data of each channel

    intensity_dv_reshape = transfer_pos(intensity_digital_value)

    # cal_result = np.array(Parallel(n_jobs=mp.cpu_count() - 1)(delayed(process_itg)(target_reshape[:, i], qe_array, tr_array) for i in range(len(target_reshape[0]))))
    # cal_result = process_itg(intensity_dv_reshape[:, 500], exposure_time)
    cal_result = np.array(Parallel(n_jobs=mp.cpu_count() - 1)(delayed(process_itg)(intensity_dv_reshape[:, i], exposure_time) for i in range(len(intensity_dv_reshape[0]))))
    t_cal = cal_result[:, 2]
    a_cal = cal_result[:, 0]
    b_cal = cal_result[:, 1]
    t_cal_map = transfer_neg(t_cal, intensity_digital_value)
    t_diff, t_diff_ref = compare(t_cal, temperature_center)
    t_diff_map = transfer_neg(t_diff, intensity_digital_value)
    t_diff_ref_map = transfer_neg(t_diff_ref, intensity_digital_value)
    save_file(t_cal_map, temperature_center, emissivity_set, 0, t_diff_map, t_diff_ref_map, result_dir)
    return 0


# def denoise(fig):
#     out_fig = gaussian_filter(fig, sigma = 1)
#     return out_fig


def compare(t_cal, temperature_center):
    t_should = np.ones(t_cal.shape) * (temperature_center + 273.15)
    t_diff = t_cal - t_should
    t_diff_ref = t_diff / t_should
    return t_diff, t_diff_ref


def save_file(t_field, temperature_center, emissivity_set, emi_field, t_diff, t_diff_ref, result_dir):
    dir_name = 'T' + str(temperature_center) + '_' + str(emissivity_set) + '_digital'
    result_dir = os.path.join('results', result_dir)
    if not os.path.exists( result_dir):
        os.mkdir(result_dir)
    if not os.path.exists(os.path.join(result_dir, dir_name)):
        os.mkdir(os.path.join(result_dir, dir_name))

    # t_field
    file_t = os.path.join(result_dir, dir_name, 't_cal_' + str(temperature_center) + '.xlsx')
    workbook_t = openpyxl.Workbook()
    worksheet_t = workbook_t.active
    plt.imshow(t_field, cmap='viridis')
    plt.colorbar()
    plt.xlabel('X_position')
    plt.ylabel('Y_position')
    plt.title('Temperature_map')
    plt.savefig(os.path.join(result_dir, dir_name, 't_cal' + '.jpg'))
    plt.clf()

    for row in t_field:
        worksheet_t.append(list(row))
    workbook_t.save(file_t)

    # t_field_denoise = denoise(t_field)
    # plt.imshow(t_field_denoise, cmap='viridis')
    # plt.colorbar()
    # plt.xlabel('X_position')
    # plt.ylabel('Y_position')
    # plt.title('Temperature_map')
    # plt.savefig(os.path.join(result_dir, dir_name, 't_cal_denoise' + '.jpg'))
    # plt.clf()

    # t_diff
    file_t = os.path.join(result_dir, dir_name, 't_diff_abs' + str(temperature_center) + '.xlsx')
    workbook_t = openpyxl.Workbook()
    worksheet_t = workbook_t.active
    plt.imshow(t_diff, cmap='viridis')
    plt.colorbar()
    plt.xlabel('X_position')
    plt.ylabel('Y_position')
    plt.title('Temperature_diff_abs')
    plt.savefig(os.path.join(result_dir, dir_name, 't_diff' + '.jpg'))
    plt.clf()

    for row in t_diff:
        worksheet_t.append(list(row))
    workbook_t.save(file_t)

    # t_diff_ref
    file_t = os.path.join(result_dir, dir_name, 't_diff_ref' + str(temperature_center) + '.xlsx')
    workbook_t = openpyxl.Workbook()
    worksheet_t = workbook_t.active
    plt.imshow(t_diff_ref, cmap='viridis')
    plt.colorbar()
    plt.xlabel('X_position')
    plt.ylabel('Y_position')
    plt.title('Temperature_diff_ref')
    plt.savefig(os.path.join(result_dir, dir_name, 't_diff_ref' + '.jpg'))
    plt.clf()

    for row in t_diff_ref:
        worksheet_t.append(list(row))
    workbook_t.save(file_t)

    # emi_field
    # file_emi = os.path.join(result_dir, dir_name, 'emi_cal_' + str(temperature_center) + '.xlsx')
    # workbook_emi = openpyxl.Workbook()
    # worksheet_emi = workbook_emi.active
    # plt.imshow(emi_field, cmap='viridis')
    # plt.colorbar()
    # plt.xlabel('X_position')
    # plt.ylabel('Y_position')
    # plt.title('Emissivity_map')
    # plt.savefig(os.path.join(result_dir, dir_name, 'emi_cal' + '.jpg'))
    # plt.clf()
    #
    # for row in emi_field:
    #     worksheet_emi.append(list(row))
    # workbook_emi.save(file_emi)


def process_itg(intensity_dv, exposure_time):
    wl = np.array([0.548, 0.586, 0.628, 0.667, 0.704, 0.743, 0.786, 0.826]) * 10 ** (-6)
    sens_factor = np.array(
        [1754.7780081330168, 4614.964564504549, 9544.836689465254, 18681.526160164747, 28448.45940189293,
         42794.15859091206, 61722.9332334226, 89214.96448715679])
    sens_factor_b = [-625203.4109383911, -1255959.3399823632, -2280428.045299191, -2896486.193421305,
                     -3511055.365513118, -2647879.5561356354, -496042.6119804597, 6119684.353094454]
    popt, conv = curve_fit(transfer_pv2dv, wl,
                           (intensity_dv * sens_factor / exposure_time + sens_factor_b),
                           bounds=((-np.inf, -np.inf, 500), (np.inf, np.inf, 3000)), maxfev=100000000)
    return popt


def image_read(data_path):
    folder_path = os.path.join('data', 'black_body_radiation', data_path)
    file_list = sorted([f for f in os.listdir(folder_path) if f.endswith(".tiff")])

    # define dimension of the intensity data
    first_file_path = os.path.join(folder_path, file_list[0])
    first_image = tifffile.imread(first_file_path)
    height, width = first_image.shape

    data = np.zeros((len(file_list), height, width), dtype=first_image.dtype)

    # read intensity data into 3D-array
    for i, file_name in enumerate(file_list):
        file_path = os.path.join(folder_path, file_name)
        image = tifffile.imread(file_path)
        data[i] = image

    # find exposure time in file name
    match = re.search(r"_([0-9]+)_", file_list[0])
    if match:
        exposure_time = float(match.group(1)) * 0.1e-3 # [micrometer]

    return data, exposure_time


def camera_read(camera_folder):
    df_qe = pd.read_excel(os.path.join(camera_folder, "CMS22010236.xlsx"), 'QE')
    df_t = pd.read_excel(os.path.join(camera_folder, "FIFO-Lens_tr.xls"))

    tr_array = np.array(df_t).transpose()
    qe_array = []
    for i in range(8):
        qe_array.append(df_qe.iloc[:, [0, 1 + i]])
    qe_array = np.array(qe_array).transpose(0, 2, 1)
    return tr_array, qe_array


def transfer_pos(input):
    return input.reshape(input.shape[0], -1)


def transfer_neg(input, target_value):
    shape_data = [target_value.shape[1], target_value.shape[2]]
    return input.reshape(shape_data)


def radiation_pv(temperature, wavelength):
    radiation_physical_value = black_body_radiation(temperature, wavelength) * emissivity_model(wavelength, 0)
    return radiation_physical_value


def black_body_radiation(temperature, wavelength):
    param1 = h * 2 * c ** 2
    param2 = h * c / k
    result_value = param1 / (wavelength ** 5) / (np.exp(param2 / (wavelength * temperature)) - 1)
    return result_value


def emissivity_model(wavelength, *parameter):
    emissivity = 0.2
    return emissivity


def transfer_pv2dv(wavelength, a, b, temperature):
    # transfer physical value into digital value
    dv = a * radiation_pv(temperature, wavelength) + b

    return dv


if 1:
    start_time = time.perf_counter()

    data_path = 'V0_bb_700'
    temperature_center = 973
    t_estimate(data_path, temperature_center)

    # data_path = 'V0_bb_750'
    # temperature_center = 750
    # t_estimate(data_path, temperature_center)
    #
    # data_path = 'V0_bb_800'
    # temperature_center = 800
    # t_estimate(data_path, temperature_center)
    #
    # data_path = 'V0_bb_850'
    # temperature_center = 850
    # t_estimate(data_path, temperature_center)
    #
    # data_path = 'V0_bb_900'
    # temperature_center = 900
    # t_estimate(data_path, temperature_center)
    #
    # data_path = 'V0_bb_950'
    # temperature_center = 950
    # t_estimate(data_path, temperature_center)
    #
    # data_path = 'V0_bb_1000'
    # temperature_center = 1000
    # t_estimate(data_path, temperature_center)
    #
    # data_path = 'V0_bb_1050'
    # temperature_center = 1050
    # t_estimate(data_path, temperature_center)
    #
    # data_path = 'V0_bb_1075'
    # temperature_center = 1075
    # t_estimate(data_path, temperature_center)
    #
    # data_path = 'V0_bb_1085'
    # temperature_center = 1085
    # t_estimate(data_path, temperature_center)

    end_time = time.perf_counter()
    print("calculation time: ", end_time - start_time, " second")