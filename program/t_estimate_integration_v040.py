import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import time
import math
import warnings
import openpyxl
import multiprocessing as mp
from scipy.constants import c, h, k
from scipy.optimize import curve_fit, least_squares
from scipy.integrate import quad
from scipy.interpolate import interp1d
from joblib import Parallel, delayed


def t_estimate_integration(result_dir, data_temperature, emissivity_set):
    currentdir = os.getcwd()
    homefolder = os.path.dirname(currentdir)
    result_dir = os.path.join('results', result_dir)
    data_name = 'T' + data_temperature + '_' + emissivity_set + '_digital'
    camera_folder = os.path.join(homefolder, 'program', 'camera_parameter')
    DF_QE = pd.read_excel(os.path.join(camera_folder, "CMS22010236.xlsx"), 'QE')
    DF_T = pd.read_excel(os.path.join(camera_folder, "FIFO-Lens_tr.xls"))

    t_melt = 1700

    tr_array = np.array(DF_T).transpose()
    qe_array = []
    for i in range(8):
        qe_array.append(DF_QE.iloc[:, [0, 1+i]])
    qe_array = np.array(qe_array).transpose(0, 2, 1)

    experiment_folder = os.path.join(homefolder, 'program', 'data', data_name)

    intensity = []
    for i in range(8):
        intensity.append(pd.read_excel(os.path.join(experiment_folder, ('digital_value_' + data_temperature + '.xlsx')), 'channel_' + str(i), header=None))
    t_ref = np.array(pd.read_excel(os.path.join(experiment_folder, 't_field_' + data_temperature + '.xlsx'), header=None))
    intensity = np.array(intensity)
    target = intensity
    t_target = t_ref
    t_map = np.zeros((len(target[0]), len(target[0, 0])))
    Ea_map = np.zeros((len(target[0]), len(target[0, 0])))
    Eb_map = np.zeros((len(target[0]), len(target[0, 0])))
    Eb_1_map = np.zeros((len(target[0]), len(target[0, 0])))
    emi_map = np.zeros((len(target[0]), len(target[0, 0])))

    # start calculation
    # parallel computing
    target_reshape = transfer_pos(target)
    print("Number of processors: ", mp.cpu_count())
    cal_result = np.array(Parallel(n_jobs=mp.cpu_count()-1)(delayed(process_itg_v030)(target_reshape[:, i], qe_array, tr_array) for i in range(len(target_reshape[0]))))
    recalculate_index = np.where(cal_result[:, 0] > t_melt)[0]
    target_recalculate = target_reshape[:, recalculate_index]
    cal_result_new = np.array(Parallel(n_jobs=mp.cpu_count()-1)(delayed(process_itg_v020)(target_recalculate[:, i], qe_array, tr_array) for i in range(len(target_recalculate[0]))))
    # 报错原因：当只有一个点温度超过2600时，cal_result_new变成一维数组，由此出现索引错误。解决方法：判断cal_result_new维度
    print(len(cal_result_new[:, 0]))
    print(len(cal_result[recalculate_index, 0]))
    # start replacing the recalculated results
    cal_result[recalculate_index, 0] = cal_result_new[:, 0]
    cal_result[recalculate_index, 1] = cal_result_new[:, 1]
    cal_result[recalculate_index, 2] = cal_result_new[:, 2]
    cal_result[recalculate_index, 3] = np.NaN

    t_map = transfer_neg(cal_result[:, 0], target)
    Ea_map = transfer_neg(cal_result[:, 1], target)
    Eb_map = transfer_neg(cal_result[:, 2], target)
    Eb_1_map = transfer_neg(cal_result[:, 3], target)

    for i in range(Ea_map.shape[0]):
        for j in range(Ea_map.shape[1]):
            emi_map[i, j] = emissivity_average_cal_v030(Ea_map[i, j], Eb_map[i, j], Eb_1_map[i, j])

    save_file(t_map, data_temperature, emissivity_set, emi_map, result_dir)


def compare(original_data, result_dir):
    # original_data = 'T1896_2_digital'
    original_data_address = os.path.join('data', original_data)

    cal_data = original_data
    cal_data_address = os.path.join('results', result_dir, cal_data)


    # read data from excel file
    for file in os.listdir(original_data_address):
        if file.endswith('.xlsx') and file.startswith('t_field'):
            df = pd.read_excel(os.path.join(original_data_address, file), header=None)
            t_target = df.to_numpy()

    for file in os.listdir(original_data_address):
        if file.endswith('.xlsx') and file.startswith('emi_field'):
            df = pd.read_excel(os.path.join(original_data_address, file), header=None)
            emi_target = df.to_numpy()

    for file in os.listdir(cal_data_address):
        if file.endswith('.xlsx') and file.startswith('t_cal'):
            df = pd.read_excel(os.path.join(cal_data_address, file), header=None)
            t_cal = df.to_numpy()

    for file in os.listdir(cal_data_address):
        if file.endswith('.xlsx') and file.startswith('emi_cal'):
            df = pd.read_excel(os.path.join(cal_data_address, file), header=None)
            emi_cal = df.to_numpy()

    t_bias = (t_target - t_cal) / t_target
    emi_bias = (emi_target - emi_cal) / emi_target

    ######### save fig
    # t_cal
    plt.imshow(t_cal, cmap='viridis')
    plt.colorbar()
    plt.xlabel('X_position')
    plt.ylabel('Y_position')
    plt.title('Temperature_cal')
    plt.savefig(os.path.join(cal_data_address, 'T_cal.jpg'))
    plt.clf()

    # emi_cal
    plt.imshow(emi_cal, cmap='viridis')
    plt.colorbar()
    plt.xlabel('X_position')
    plt.ylabel('Y_position')
    plt.title('emissivity_calculate')
    plt.savefig(os.path.join(cal_data_address, 'emi_cal.jpg'))
    plt.clf()

    # bias
    file_t = os.path.join(cal_data_address, 't_bias.xlsx')
    workbook_t = openpyxl.Workbook()
    worksheet_t = workbook_t.active

    for row in t_bias:
        worksheet_t.append(list(row))
    workbook_t.save(file_t)

    plt.imshow(t_bias, cmap='viridis')
    plt.colorbar()
    plt.xlabel('X_position')
    plt.ylabel('Y_position')
    plt.title('Temperature_bias_rel')
    plt.savefig(os.path.join(cal_data_address, 'T_bias.jpg'))
    plt.clf()

    file_emi = os.path.join(cal_data_address, 'emi_bias.xlsx')
    workbook_emi = openpyxl.Workbook()
    worksheet_emi = workbook_emi.active

    for row in emi_bias:
        worksheet_emi.append(list(row))
    workbook_emi.save(file_emi)

    plt.imshow(emi_bias, cmap='viridis')
    plt.colorbar()
    plt.xlabel('X_position')
    plt.ylabel('Y_position')
    plt.title('emissivity_bias_rel')
    plt.savefig(os.path.join(cal_data_address, 'emi_bias.jpg'))
    plt.clf()


# transfer 3d array into a 2d array for parallel computing
def transfer_pos(input):
    return input.reshape(input.shape[0], -1)


def transfer_neg(input, target_value):
    shape_data = [target_value.shape[1], target_value.shape[2]]
    return input.reshape(shape_data)


def emissivity_average_cal_v030(a, b, b_1):
    wl0 = 500
    wl1 = 1000
    wavelength = np.arange(wl0, wl1) * 10**(-9)
    if np.isnan(b_1):
        emissivity = np.average([emissivity_model_v020(wl, a, b) for wl in wavelength])
    else:
        emissivity = np.average([emissivity_model_v030(wl, a, b, b_1) for wl in wavelength])
    return emissivity


def lin_interpolation(x, x0, x1, y0, y1):
    return y0+(y1-y0)*(x-x0)/(x1-x0)


def black_body_radiation(temperature, wavelength):
    param1 = h * 2 * c ** 2
    param2 = h * c / k
    return param1/(wavelength**5)/(np.exp(param2/(wavelength*temperature))-1)


def integration_v020(wl, f_array, qe_array, a, b, t):
    wl0 = 0.5 * 10 ** (-6)
    wl1 = 1 * 10 ** (-6)
    f_i = len(f_array[0,f_array[0,:]*10**(-9)<=wl]) - 1
    qe_i = len(qe_array[0,qe_array[0,:]*10**(-9)<=wl]) - 1
    f = lin_interpolation(wl*10**9,f_array[0,f_i-1],f_array[0,f_i],f_array[1,f_i-1],f_array[1,f_i])
    qe = lin_interpolation(wl*10**9,qe_array[0,qe_i-1],qe_array[0,qe_i],qe_array[1,qe_i-1],qe_array[1,qe_i])
    # result = f*(a-b*(wl-wl0)/(wl1-wl0))*qe*black_body_radiation(t,wl)*200
    result = f * emissivity_model_v020(wl, a, b) * qe * black_body_radiation(t, wl) * 200
    return result


def integration_v030(wl, f_array, qe_array, a, b, b_1, t):
    wl0 = 0.5 * 10 ** (-6)
    wl1 = 1 * 10 ** (-6)
    f_i = len(f_array[0,f_array[0,:]*10**(-9)<=wl]) - 1
    qe_i = len(qe_array[0,qe_array[0,:]*10**(-9)<=wl]) - 1
    f = lin_interpolation(wl*10**9,f_array[0,f_i-1],f_array[0,f_i],f_array[1,f_i-1],f_array[1,f_i])
    qe = lin_interpolation(wl*10**9,qe_array[0,qe_i-1],qe_array[0,qe_i],qe_array[1,qe_i-1],qe_array[1,qe_i])
    # result = f*(a-b*(wl-wl0)/(wl1-wl0))*qe*black_body_radiation(t,wl)*200
    result = f * emissivity_model_v030(wl, a, b, b_1) * qe * black_body_radiation(t, wl) * 200
    return result


def emissivity_model_v030(wl, a, b, b_1):
    # lin square emi = a + b * wl**2
    wl0 = 0.5 * 10 ** (-6)
    wl1 = 1 * 10 ** (-6)
    wl_rel = (wl - wl0) / (wl1 - wl0)
    emissivity = a * wl_rel**2 + b * wl_rel + b_1
    return max(0, min(emissivity, 1))


def emissivity_model_v020(wl, a, b):
    # exp emi = exp(-a - b * wl)
    wl0 = 0.5 * 10 ** (-6)
    wl1 = 1 * 10 ** (-6)
    wl_rel = (wl - wl0) / (wl1 - wl0)
    emissivity = math.exp(-a - b * wl_rel)
    return max(0, min(emissivity, 1))


def process_itg_v030(intensity_array, qe_array, tr_array):
    wl0 = 0.5 * 10 ** (-6)
    wl1 = 1 * 10 ** (-6)

    def integration_solve(qe, a, b, b_1, t):
        result_f = []
        for i in range(8):
            funct = quad(integration_v030, wl0, wl1, args=(tr_array, qe[i], a, b, b_1, t), epsabs = 1e-2, limit = 5)[0]
            result_f.append(funct)
        return np.array(result_f)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        popt, cov = curve_fit(integration_solve, qe_array, intensity_array, bounds=((-3, -3, -1, 1200), (3, 3, 1, 2000)), maxfev= 100000)
    return popt[3], popt[0], popt[1], popt[2]


def process_itg_v020(intensity_array, qe_array, tr_array):
    wl0 = 0.5 * 10 ** (-6)
    wl1 = 1 * 10 ** (-6)

    def integration_solve(qe, a, b, t):
        result_f = []
        for i in range(8):
            funct = quad(integration_v020, wl0, wl1, args=(tr_array, qe[i], a, b, t), epsabs = 1e-2, limit = 5)[0]
            result_f.append(funct)
        return np.array(result_f)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        popt, cov = curve_fit(integration_solve, qe_array, intensity_array, bounds=((-1, -1, 1200), (1, 1, 2000)), maxfev= 100000)
    return popt[2], popt[0], popt[1]


def save_file(t_field, temperature_center, emissivity_set, emi_field, result_dir):
    dir_name = 'T' + str(temperature_center) + '_' + str(emissivity_set) + '_digital'
    if not os.path.exists(result_dir):
        os.mkdir(result_dir)
    if not os.path.exists(os.path.join(result_dir, dir_name)):
        os.mkdir(os.path.join(result_dir, dir_name))

    # t_field
    file_t = os.path.join(result_dir, dir_name, 't_cal_' + str(temperature_center) + '.xlsx')
    workbook_t = openpyxl.Workbook()
    worksheet_t = workbook_t.active
    plt.imshow(t_field, cmap='viridis')
    plt.colorbar()
    plt.xlabel('X_position')
    plt.ylabel('Y_position')
    plt.title('Temperature_map')
    plt.savefig(os.path.join(result_dir, dir_name, 't_cal' + '.jpg'))
    plt.clf()

    for row in t_field:
        worksheet_t.append(list(row))
    workbook_t.save(file_t)

    # emi_field
    file_emi = os.path.join(result_dir, dir_name, 'emi_cal_' + str(temperature_center) + '.xlsx')
    workbook_emi = openpyxl.Workbook()
    worksheet_emi = workbook_emi.active
    plt.imshow(emi_field, cmap='viridis')
    plt.colorbar()
    plt.xlabel('X_position')
    plt.ylabel('Y_position')
    plt.title('Emissivity_map')
    plt.savefig(os.path.join(result_dir, dir_name, 'emi_cal' + '.jpg'))
    plt.clf()

    for row in emi_field:
        worksheet_emi.append(list(row))
    workbook_emi.save(file_emi)


if 1:
    start_time = time.perf_counter()

    result_dir = 'result_v040_lin_square_exp'
    data_temperature = '1900'
    emissivity_set = '5'
    data_name = 'T' + data_temperature + '_' + emissivity_set + '_digital'
    t_estimate_integration(result_dir, data_temperature, emissivity_set)
    compare(data_name, result_dir)
    print(data_name, 'finished')

    end_time = time.perf_counter()
    print("calculation time: ", end_time - start_time, " second")